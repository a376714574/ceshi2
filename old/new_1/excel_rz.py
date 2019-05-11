'''
读写excel
'''
from openpyxl import load_workbook

class Case:

      def __init__(self):
          self.id=None
          self.title=None
          self.url=None
          self.data=None
          self.mode=None
          self.Exq=None

class excel:

    def __init__(self,excel_name=None,sheet=None):
        self.excel_name_z=excel_name
        self.excel_name = load_workbook(excel_name)
        self.sheet = self.excel_name[sheet]

    def conf(self):
        sw=self.sheet.cell(2,2).value
        self.excel_name.close()
        return sw

    def excel_read(self):
        excelread_list = []

        for i in range(2,self.sheet.max_row+1):
            case = Case()
            case.id=self.sheet.cell(i,1).value
            case.title=self.sheet.cell(i,2).value
            case.url=self.sheet.cell(i,3).value
            case.data=self.sheet.cell(i,4).value
            case.mode=self.sheet.cell(i,5).value
            case.Exq=self.sheet.cell(i,6).value
            excelread_list.append(case)
        self.excel_name.close()
        return excelread_list

    def excel_write(self,id,actual_results,test_results):
        id += 1
        for x in range(2,self.sheet.max_row+1):
            self.sheet.cell(id,7).value=actual_results
            self.sheet.cell(id,8).value=test_results
            self.excel_name.save(self.excel_name_z)
        self.excel_name.close()
