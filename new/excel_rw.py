from openpyxl import load_workbook

class Case:

    def __init__(self):
        self.id = None
        self.title = None
        self.url = None
        self.data = None
        self.mode = None
        self.exp = None
        self.act = None

class excel:

    def __init__(self,ex_path,sheet):

        self.excel_path=load_workbook(ex_path)
        self.sheet=self.excel_path[sheet]


    def excel_r(self):

        ex_list = []

        for i in range(2,self.sheet.max_row+1):
            case = Case()
            case.id = self.sheet.cell(i,1).value
            case.title = self.sheet.cell(i,2).value
            case.url = self.sheet.cell(i,3).value
            case.data = self.sheet.cell(i,4).value
            case.mode = self.sheet.cell(i,5).value
            case.exp = self.sheet.cell(i,6).value
            case.act=self.sheet.cell(i,7).value
            ex_list.append(case)
        return ex_list

    def excel_w(self):
        pass

    def excel_close(self):
        self.excel_path.close()



