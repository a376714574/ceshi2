'''
----------------------------------------------------------------------
基础知识

字符串种类  单引号 双引号 三引号 三引号是可以写多行的
单行注释 快捷键ctel+/                        Python识别大小写
标识符:只能包含数字字母下划线 不能以数字开头 不能用关键 第三方模块 源

import keyword
print(keyword.kwlist) 能查寻到所有的关键字None True and as等

a='a'\  两行的数据拼接                   a='a' 'b'  一行里面不需要其他东西也可以拼接
'b'
int(a())   str(a())  强制转换

a,b,c=1 这样也可以赋值                 a=b=c=4 这样也可以赋值
----------------------------------------------------------------------
常用函数
type()                    查询数据类型
len()                     查询长度 一个汉字代表一个长度
round(a*0.8,2)            强制定义小数点后面的位数
input('请输入程序的成绩')    输入函数返回字符串
----------------------------------------------------------------------
数据类型
int str 元祖 列表 字典 浮点                    300.0  只要有小数点就是浮点数
布尔值 True 1/非空  False 0/空数据t

元祖 tuple--------------
()用它括起来的就是一个元祖 元祖不能为空 (,)这样是可以的
(1) 这就是一个元祖  数组里面的元素是不可以修改的
元祖里面可以放入任意类型 str int fo 布尔值 元祖（自己里面放自己）.....

列表 List[]------------------------------------------
可变数据类型可以为空
s=[4,5,6]
L=L+s               列表拼接
L.insert(1,"aaaa")  添加到指定位置 那个位置的元素往后排
L.extend(S)         拼接两个列表
L.append("111")     添加一个元素到列表的最后面
L[1]="asd"          指定位置替换
L.pop(2)            删除指定元素 不写默认删除最后一个
L.clear()           清空列表
L.remove('hello')   会删除第一次遇到的这个元素后面的不会删除
range(1,6,1)        元素索引
L.reverse()         列表反转
L.sort()            列表排序仅限数字
L.sort(reverse=True)先排序后反转
L.copy()            复制出来创建一个新的对象 不受原来的对象影响
L.sort()            数字排序从小到大
L.sort(reverse=True)就是排序后再反转
print(a[4][3][::-1])
嵌套取值
a=(1,['124',1],1.211,True)
print(a[1][1])

L=[100,'hello',(1,2,3),['python',0.02]]
for a in range(0,4): 会把 0 1 2 3依次赋值给a
    print(L[a])       这里面那都是步长输出的结果都是数字 然后都输出出来

L=[100,'hello',(1,2,3),['python',0.02]]
for a in range(len(L)-1,-1,-1): 倒序输出 len(L)是4  结果就是3 -1 -1 从三开始到-1 位置
    print(L[a])

字典  dict(key value)--------------------------
key  不可变数据类型 int float str boolean tuple 这些是不可变的 这些可以使用
value 可以改变的不可改变 所有都可以使用 乱序输出没有规律 如果key的值如果重复 前面的V会被替换掉
a={1:'我',True:'爱','age':22}  这里的key一个是1 另一个是True 但是他们的是一样的所以被覆盖了

print(a["age"])               输入key来取值V
d['name']='dsaas'             修改对应key  V的值 如果key不存在那么会创建一个出来
d.pop('name')                 通过key去删除这个k V
d.popitem()                   随机删除一个k v
d.copy()                      复制
d.items()                     全部打印
d.get('name')                 根据key 来返回V

字典取值
d={'1':'2','a1':'2','b1':'2','c1':'2',}
for a in d.items(): #也可以写b.keys()那么取得就是key  b.V取出来的就是V
                                   d.items()类型是元祖会把k V都取出来给a
for a,c in d.items(): 这样写 a取的就是key   c取的就是V  写出了a，c a就是接收key的值 c就是接收V的值

json和字典的区别
json  object    array      string             number        true   false  null
dict  dict   list,tule   str,unicode      int long float    True   False  None

json 是一种数据格式 最初是jq使用的语法前端和后端传输数据使用的  json里面只有null没有None
python里面有None没有null 所以空数据会显示None
json格式：{"1":"12"} 所有的字符串都必须是双引号   数据是空的时候使用的null

eval和json的区别

eval(a)#根据python的数据类型来做转换 但是"regname": "null" python里面没有null只有None所以会报错
a='{"mobilephone": "","pwd": "123456","regname": "null"}'
b=json.loads(a) #转换字典 可以用key去取出他的值
-------------------------------------------------------
字符串操作
s='y2yyyyyy2y'

print('你'+'好')       字符串拼接的方式
s.strip('s')          去掉头和尾的内容可以是空格
s.capitalize()        首字母大写
s.swapcase()          大小写呼唤
s.count('y',2,8)      检查y出现的次数 2,8这个是范围不写默认是全部  像是数组一样0,1,2
s.find('y',2,8)       找到第一个y出现的位置 和数组一样0,1,2 如果没有则返回-1 不写默认是全部
s.index('H',2,8)      和find一样只是没找到的话会报错 不写默认是全部
s.islower()           判断是否都是小写 是返回T 否返回F
s.isupper()           判断是否都是大写 和上面一样
s.upper()             强制全部转换成为大写
s.isdecimal()         判断是否都是数字类型只能是int  但只能判断十进制
s.lower()             强制全部变成小写
'@'.join(s)           每一个元素之间都用@分隔开 @可以填成为任何
s.replace('p','@',1)  把p替换成为@ 后面是替换次数 写500也可以不会报错
s.split('y',1)        切割字符串变成列表 前面是切割内容 后面是次数 被切割的掉内容会变成一个空的元素 所有的y都会被删掉变成一个空元素
print(a,b,sep="ddd",end='\n') sep 会把那两个字符串隔开ddd是隔开的隔开的内容 end换行

索引 index(字符串取值)
s='我爱北京天安门'

s[0]           正向取值       s[-1]          反向取值
s[::]          正向全取值     s[::-1]        反向全部取值
   0 1 2 3 4 5 6
s='我爱北京天安门'
s[4:6:1]             4:6 是范围 1是步长 步长就是4 5 6的取值 如果是2 那么就是4 6 8这样的不写的话默认就是1
                         取出的结果是"天安"  4 5 6应该是"天安门" 但是取头不取尾 所以取的就是4 5
s[::]    正无穷~负无穷取值
s[::-1]  负无穷~正无穷取值
s[-1::-1]负无穷~正无穷取值
s[-1::-1]索引头是-1 索引尾是负无穷

字符串制作翻译表
s='pytHonpyyppp'

intab = "aeiou"
outtab = "12345"    #制作翻译表
trantab = str.maketrans(intab, outtab)  #把翻译表放进去
str = "this is string example....wow!!!"#要翻译的语句
str.translate(trantab)                  #进行翻译
-------------------------------------------------------
格式化输出
a=15.1    b="ad"    c=0.15

第一种格式化输出%s 字符串    %d 数字    %f  浮点数
print('%s,我是名字%d,我是薪资%f'%(a,b,c))
print('%.2f'%(a))  %.2f 保留小数点后面两位3f就是三位 负数也可以
print('名字:{},薪资:{},薪资:{}'.format(a,b,c)) 这是另一种表达方式
---------------------------------------------------------------------
函数的语法
print()  len()这些都是函数 只不过都写好了 直接用就好

def 函数名(参数1，参数2，参数3... ...)
def add(a,b):            参数
    return               返回值 必须要返回数据才能使用
    return ['1','2'],1   如果返回多个值会把他们装进元祖
add(1,2)                 调用函数并且传参
add(b=1,a=2)             这样的写法可以无视顺序 但如果没有指定会按照函数定义的顺序执行
b=a(),4                  这也是一种写法
a,b=add()                返回值是[1,2],2有两个值 所以用a b来接收如果用一个 就会变成一个元祖

def add(a,b=9)           b=9是默认参数 函数定义了几个参数就要穿几个 但是如果有默认参数 传不传都行 如果没传他就默认使用默认参数
                         默认参数可以有多个但是必须放在坐后面比如(a,b,c=1，d=2，d=3)最后面 最后面一定是默认参数

局部变量/全局变量--------------------------
def
 H():
    a="python"            a是局部变量 只能在函数内出现效果 如果在这个函数内调用a变量优先调用局部变量

def a():
    global b              global全局变量的意思 也就是说把b变成全局变量 虽然她本身是一个局部变量 这样一来就变成全局变量了
    b='python'

函数中的动态参数--------------------------
def a(*args):    *args动态参数可以传多个值 会把它们放到一个元祖里              def a(*a):
     b=0                                                                        print(a)
     for c in args:           取值                                              t = [1, 2, 3, 4, 5]
          b+=c               求和的一个方法                               a(*t)  结果是(1,2,3,4,5)    *会解包拆开list
     print(b)                                                           a(t)   结果是([1,2,3,4,5],)没有*会把它直接放到元祖里
a()可以是空值传输

def a(**kwargs):      **kwargs  动态参数可以穿多个参数会变成字典
    print(kwargs)

a(a=1,b=2)    这样传参结果就是{a:1,b:2}
c={"222":"1","211":"2"} 这样传输也可以
a(**c)
----------------------------------------------------------------------
运算符
算术运算符+ - / % //（地板除不要小数点直接得到答案）
print(10/3)  结果是3.333333333
print(10%3)  结果是1 取余数      print(9·%3)如果整除就是0 可以用于布尔值来操作
print(10//3) 地板除结果是3 因为3乘以3后就是9不能再乘了 print(100//3) 结果就是33

逻辑运算符
and 和       or 或         and not         or not       运算级别 not>and>or
print(a and b)      两边都是T才是T 两个都是F 就是F 一个是F也是F
print(a<0 or b==0)  有一个是T结果是T

成员运算符   in 在里面          not in 不再里面
a='python'
print('p' in a) p是否在a的里面结果是T 因为在里面了
----------------------------------------------------------
with上下文管理器
防止忘记关闭 有了上下文管理器 他会自动关闭

with open("123.text","w+",encoding="utf-8") as file:
     file.write("123")
print(file.closed) #检查是否关闭T是关闭 F是未关闭
file.write("123")  #已经关掉后就不能再写入了
----------------------------------------------------------
异常处理
异常处理是为了防止程序运行停止 抛出一个异常 程序继续运行

错误类型
NameError
IndexError
KeyError
OverflowError 数值运算超出最大限制
FloatingPointError  浮点运算错误
ArithmeticError   所有数值计算错误的基类
StandardError     所有的内建标准异常的基类
GeneratorExit     生成器发生异常来通知推出
Stoplieration    迭代器没有更多之
Exception        常规错误的基类
Keyboardinterrupt  用户终端执行通常是输入'C
SystemExit        解  释  器请求推出
BaseException     所有异常的基类
ZeroDivisionError 除(或取模)零(所有数据类型)
AssertionError    断言语句失败
AttributeError    对象没有这个属性
EOFError          没有内建输入 到达EOF标记
EnvironmentError  操作系统错误的基类
IOError           输入/输出操作失败
OSError           操作系统错误
WindowsError      系统调用失败
importError       导入模块/对象失败
LookupError       无效数据查询的基类
indexError        序列中没有此索引(index)
KeyError          映射中没有这个key
NameError         未声明/初始化对象(没有属性)
SyntaxError       Python语法错误
MemortError       内存溢出错误
indentationError  缩进错误
TabError          空格和Tab混淆

try:                  #监控
  s=[1,2,3]           #被检控的内容
  print(s[4])         #虽然下面可以写多种错误 但是他只是会找到第一条后面的错误就不处理了
  print(s)            #不会执行因为上一行已经错误
except (NameError,IndexError) as e:#监控的错误类型 as取别名 必须取别名 如果没监控的错误类型 出现了错误就会报错
  print("出错了")      #如果出现了异常这里就会运行了 如果没有就不会运行这里
  print(e)            #输出错误类型
#try...except....他的作用就是拦截异常 如果没有他 程序就不会执行了 但是有了他们 他们会拦截住后面的继续运行
else:
  print("else")       #如果try是正确的 就会运行else   如果try报错了else就不执行
finally:
     print(s)         #实际上有没有finally都可以执行后面的 因为上面已经把异常拦截住了
--------------------------------------------------------------------------------
Logging 日志:

log也有等级 debug info warning error critical/fatal
如果没有定义控制器 就会使用默认控制器叫做root 默认输出到控制台 默认输出到warning级别包括它本身

import logging

a=logging.getLogger("啦啦啦") #创建一个收集器名字叫做"啦啦啦"
a.setLevel("DEBUG")    #设置收集级别

上面是创造输入器 下面是创造输出器--------------------------------------------------
PS:如果收集器只收集3级别以下的 那么输出器即使全部输出也没有只有3级别以下的内容---------------------------------

fmt=a.Formatter('%(asctime)s-%(filename)s-%(levelname)s-日志信息:%(message)s') #选择你想要的输出格式

ch=a.StreamHandler()   #创建一个输出器 控制台
ch.setLevel("DEBUG")   #设置控制台输出级别
ch.setFormatter(fmt)   #定义输出格式

b.addHandler(ch)       #把输入器和输出器拼接到一起

b.debug("1")           #这一块是定义 日志等级 这里创建完后输入器和输出器才能调用
b.info("2")
b.warning("3")
b.error("4")
b.critical("5")
----------------------------------------------------------------------------------------
pip install openpyxl   需要安装第三方模块 对Excel文件进行操作
要选择xlsx 不要选择xls结尾的 Excel文件

from openpyxl import workbook

a=workbook.Workbook()
a.create_sheet("2124")                 #创建表单 起名字叫做2124
a.save(r"C:\Users\111111111.xlsx")     #保存在这个路径的文件里面 如果没有就创建一个


from openpyxl import load_workbook

a=load_workbook(r"C:\Users\111111111.xlsx")    #打开工作簿
b=a['Sheet']                       打开工作簿
c=eval(b.cell(3,2).value)          定位单元格 用坐标的方式 eval函数可以把里面的对应类型读取出来出字符串类型
c=cell(3,2).value="123"            没有写Value是取值 写了就是赋值
a.save(r"C:\Users\111111111.xlsx") 记得保存输入保存路径 要和上面对应 不然就会创建一个新的文件夹进行保存
print(c)                           输出取到的值
a.close()                          记得关闭当然如果用了with就不需要了他会自动关闭的

eval是一个转换成python的函数 比如一个字符串的长度是7 那么使用eval后他的长度变成5

for i in range(1,b.max_row+1):
    for x in range(1,b.max_column+1):
        if b.cell(i,x).value:    #这一步操作就是把none的数据排除出去
           print(b.cell(i,x).value)      #就算什么都不写答案是F的数据就不往下走了

print(b.max_row)                         #空也是F
print(b.max_column)
最大行和最大列 有数据的
------------------------------------------------------------------------------------
pip install requests

python比java简单的地方就是有非常强大的第三方库 几乎不需要取些什么 类似拼装一样品号就好了
java的底层很多都需要手动去写

request 是客户端向服务器发送的一个请求 包含请求头 请求地址 请求参数 http协议版本
response 是服务器向客户端发送一个响应 包含 响应头 响应报文 状态码等等
常见的状态码 404 200 500 301 302 502 400 所有的状态码都可以在网站上看到
200表示请求成功  但仅仅只是 发送消息和返回消息成功 就算失败他也是返回了消息
404表示找不到页面 url地址错误或者找不到

500 服务器遇到了不知道如何处理的情况 服务器的问题5开头的大部分都是服务器的问题
500 502网关错误
304 静态资源png jpg gif css js各种静态资源 在第一次请求的时候就会缓存到本地
如果这个时候再次请求的时候 并不需要向服务器去要 而是已经缓存到了客户端
所以这个时候就是304 没有更改 第一 本地就有 第二服务器并没有修改 就是这个意思
但如果清空缓存了 就不会出现这个了

import requests  需要导入requests模块

url="http://www.lemfix.com"

a=requests.get(url)                     requests.post(url)
# 发送一个get请求                        # 发送一个post请求
requests发送一个get或者post请求到服务器并且可以接收服务器的返回信息 但是他只是会返回状态吗get是200 post是302

a.text    响应信息 会把整个HTML前端页面代码返回回来post返回的更有格式 get 把所有信息变成一行
a.headers 响应头 应该是服务器返回的信息 get返回的信息 比post要多
a.status_code  状态码
a.cookies      响应的cookies

url="http:/py/47.107.168.87:8080/futureloan/mvc/api/member/login"
param_2={"mobilephone":"18688773467","pwd":"123456"}  这个类型是字典 如果不行就变成json
requests.post(url,params=param_2) 可以参数进去 比如账号密码 比如金钱数量 比如时间

a={"1":"2","1":"2","1":"2","1":None}      这个是一个字典字典的空是None 字典什么引号都行
b='{"1":"2","1":"2","1":"2","1":null}'    这是一个json字符串他的空就是null json里面必须是双引号

import json
json.loads(b)   强制转换json
---------------------------------------------------------------------------------------------
封装:
多个方法放到一个类里面 就是封装

class phohe:

      def a(self):
          pass

      def b(self):              这就是封装 多个方法放到一个类里面
          pass

      def c(self):
          pass

继承 拓展 重写

继承
from WEEK.beijichengde import *  前面的路径要精确到模块  后面的写上要继承的类名 也可以写*这样里面的类随便继承

class saber(asd):                这里面写上要继承的类名
     def pay(self):              如果父类里面也有一个叫做pay的方法 继承后再写一个同样名字的 方法就是重写了
          print("可以支付")        继承后 子类拥有父类的一切方法和属性

如果你继承的父类里面有初始化函数 那么必须传参 但如果子类重写了初始化函数 就传子类的初始化就好了
------------------------------------------------------------------------------------------
类与对象                                            类具有所有的属性和方法
万物皆对象                                          对象具有类的所有属性和方法
对象都是具体某个类里面的

创建对象   类名():                                             类的语法
                                                              class 类名：
继承 封装 重写---》语言特性 调代码的复用性 可重写性
函数和方法的区别就是差一个self 方法有对象 函数没有

对象 可以调用  类方法 静态方法  属性  对象方法       a().pr_self
类   可以调用  类方法 静态方法  属性                a.pr_self 类的调用方式

class asd:                      类

    a = 1                       属性

    def __init__(self,sex):    初始化方法 这个是python自带的不是我们定义的方法
          self.sex=sex          如果属性和初始化方法里面的属性名重复 会被复写 self.a也就代表着这是一个对象属性
          asd.a=1               创建了一个类方法
          asd().a=1               创建了一个对象属性

     def add_001(self):
          print("add_001")

     def add(self):     他是对象方法可以调用 对象方法 类方法 静态方法 属性 一切都可以万物皆对象
          self.add_001()
          self.dsaasd()
          self.self_1()
          print(self.a)

     @classmethod
     def jlk(cls):      他是类方法  cls代表了类本身调用它会把类传给他 所以他可以使用所有东西
         cls.self_1()   可以调用静态函数 类属性 类方法 不能调用对象
         print(cls.a)
         cls.dsaasd()

     @staticmethod
     def self_1():      他是静态方法 他既没有对象self 也没有类cls所以他不能直接调用其他对象 类 属性等
         print("dsasd")  但是通过创建对象就可以了asd().a这样 但这样也不是他自身的对象做到的

asd().add()

print(id(t1))        如果两个内存地址一直那么就是一个对象
print(id(t2))

asd.self_1()  调用对象的时候需要把一个对象传输进去 但是他是类不是对象
asd.self_1(N) 传输一个对象进去就可以了 asd().self_1()这样也可以 他们一个把N 另一个把asd()当作对象传给self
------------------------------------------------------------------------------------------------------------
文件处理:

a=open("../ceshi/week3_41.txt",encoding="utf-8") 读取某个内容 后面是编码
                               ../先返回一级 然后自进入ceshi最后在进入week
a=open(r'F:\vvs\123456.txt')     大小写都行  r编译里面的/ :等特殊字符大小写都可以\t \r \n
                                         等等在路径中出现这是错误的 必须\\或者在前面加上r
a.write("dsadadas")       输入内容不读直接写 会写在开头 然后覆盖前面的内容
                                    先读再写 会直接写到后面

r 只读    r+ 读写
w 写      w+ 读写
a 追加    a+ 读追加
rb rb+ wb wb+ ab ab+ 文件流的形式去写入文件的时候需要这样 binary

三种编码 utf-8 gbk unicode

a=open("week3_41.txt","r+",encoding="utf-8")如果换成w会变成清空内容后写入  如果文件不存在他会创造这个文件
a.read(10)                                   读取10个长度的内容
a.seek(2,0)                      相对于头部移两个位置第一个参数是头部的位置 根据编码方式不一样 需要考虑 因为中文
                                 是三个位置如果偏移两个所以就会报错
a.write("啦啦啦")                 写入内容但是上面必须给予权限比如w或者a+之类的
a.seek(0.0)                      移动光标 如果不把光标移动到开头是不可以读取出来的
a.readline()                     只显示第一行
a.readlines()                    读取所有行 类型变换成为列表
a.close()                        一定要关闭

asd=open("week3_411.txt",mode="r",encoding="utf-8").readline()
                                                    #读取第一行
for qwe in asd:
    qwe=qwe.strip("\n")   每读取一个元素 换行一次
    data=qwe.split(",")   每个元素之间用,分开
    print(data)
------------------------------------------------------------------------------------
路径处理
相对路径 绝对路径
绝对路径open("c:/wed/b")是一个准确的路径无论放在那里都一样
相对路径比如两个文件都在wed下面一个叫a一个叫b  那么在a文件里面调用b直接写b就行
import os

os.getcwd() 查看文件当前的位置

os.path.realpath(__file__)       输出文件当前路径 __file__指的是文件本身
os.path.realpath('week3_4')      无论这个文件在哪里都可以返回路径
a=os.path.realpath(__file__)      __file__指的是文件本身 给出的路径比上面更具体
a=os.path.realpath('week3_4')    返回一个文件的路径无论文件在哪里     print(__file__)这样也可以
os.path.basename(__file__)       输出文件名
os.path.split(a)                 他会把返回的路径切割成两个元素 一个是路径 一个是文件名变成一个元祖
os.mkdir(b[0]+'\python')         创建一个文件夹 前面是路径后面是名字
os.path.join(b[0],'python125')   join拼接路径 他会创造一个python125文件夹出来

a=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# os.path.abspath(__file__) 显示出绝对路径C:\Users\Administrator\PycharmProjects\python_15\WEEK\week3.py
# os.path.dirname显示当前文件的文件夹所以他的路径只显示到
# C:\Users\Administrator\PycharmProjects\python_15\WEEK 路径只显示到他的上一层文件夹 如果
dsss=os.path.join(a,"WEEK1","asd.test")  #可以进行自由拼接
-------------------------------------------------------------------------------------
import
导入一个模块或者一个类
用法:import 模块名
import os                        os是一个系统环境模块里面很多函数
import WEEK.week3_42222 as t     给这个模块去一个简短的名字这样比较方便使用  .py结尾的文件都可以叫做python模块
import WEEK.week3_42222          前面是路径后面是文件
C:\Python34\Lib                  python所有的模块都安装在这个路径下
from...import...                 from后面是路径 import后面是模块名 另一种from具体到模块名 import后面写具体的类名
from WEEK.week3_42222 import *   如果前面精确到了模块名 那么*就是这个模块下的所有类 如果前面只是路径 *就是这个路径下的所有模块
from WEEK.week3_42222 import addtdsx as add  昵称
from homework_0413.common.constants    import                case_file
             导入到模块名                     一般这个时候导入的都是类 但是这个导入的是一个属性 这个模块里面没有类
---------------------------------------------------------------
while循环

while a<5:后面是布尔值所以T F 空 非空 1 0都可以
   a+=1
   print("1")       break停止循环  continue 跳过本次循环 他们只会出现在while循环里面

for循环
L=["1","2","3","4","5",]
for a in L:  把L里面的每一个元素都给a
    print(a)所以他会输出L里面的所有元素

嵌套for循环
for a in s:
    print(a)
    for b in a:
        print(b)
---------------------------------------------------------------
'''


